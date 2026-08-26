"""xlsx reading for statement imports, using openpyxl.

``app/src/main/java/com/deepak/umber/io/XlsxReader.kt`` hand-rolls a ``ZipInputStream`` + SAX
reader to avoid shipping Apache POI in an APK with no network permission. The server has no such
size constraint, so this module uses ``openpyxl`` (already in requirements.txt) instead — but it
has to reproduce two behaviors of the Kotlin reader exactly, because they change *which* cells the
rest of the pipeline sees:

1. **Sheet selection.** The Kotlin reader does not read "the first tab" — it reads whichever
   worksheet part in the zip (``xl/worksheets/sheetN.xml``) sorts lexicographically first by
   filename (``sheet1.xml`` before ``sheet10.xml`` before ``sheet2.xml``). That is not always the
   same sheet as tab position 0: a workbook whose tabs were reordered in Excel keeps its original
   part filenames, so the lexicographically-first *file* and the first *tab* can differ.
   ``openpyxl``'s ``wb.worksheets`` is ordered by tab position (as declared in
   ``xl/workbook.xml``), so this module inspects the zip's own part names directly — via
   ``xl/workbook.xml`` and ``xl/_rels/workbook.xml.rels`` — to resolve the same part-name-order
   winner back to a sheet name, then asks openpyxl for that specific sheet.

2. **Raw cell text for numeric/date cells.** The Kotlin reader takes whatever is in a cell's
   ``<v>`` verbatim for any non-shared-string cell — including a cell Excel displays as a date,
   which is stored as a bare day-count integer (e.g. ``45786``) with the "date" applied only as
   display formatting. ``DateParse.date()`` on the Kotlin side relies on exactly this: its
   Excel-serial fallback exists *because* a real date-typed xlsx cell arrives as a bare number, not
   a formatted date string. openpyxl instead auto-converts a date-formatted numeric cell straight
   into a Python ``date``/``datetime`` object at load time, before this module ever sees it, so a
   naive ``str()`` of that value would produce something like "2026-05-08 00:00:00" rather than
   "45786" — a value that looks superficially date-shaped but is not what the phone would have
   parsed from the same file, and would go byte-for-byte out of parity in any format string with a
   trailing time component. This module reverses that conversion (via
   ``openpyxl.utils.datetime.to_excel``) back into the same serial-number string the Kotlin reader
   would have read from the XML, so the same Excel-serial fallback path in ``dates.py`` fires
   identically on both sides.
"""

from __future__ import annotations

import datetime as dt
import zipfile
from io import BytesIO
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.utils.datetime import to_excel

_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _lexicographically_first_sheet_name(zf: zipfile.ZipFile) -> str | None:
    """Resolves the zip's lexicographically-first worksheet part back to a sheet name.

    Mirrors XlsxReader.kt's ``name.startsWith("xl/worksheets/sheet") && name.endsWith(".xml")``
    check exactly (a substring test, not a strict "sheetN.xml" pattern) followed by keeping the
    smallest name under plain string ordering.
    """
    candidates = [
        n for n in zf.namelist() if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")
    ]
    if not candidates:
        return None
    # The zip entry name to match against, e.g. "xl/worksheets/sheet1.xml".
    target_part = min(candidates)

    try:
        rels_xml = zf.read("xl/_rels/workbook.xml.rels")
    except KeyError:
        return None
    rid_to_target: dict[str, str] = {}
    for rel in ET.fromstring(rels_xml):
        rid = rel.get("Id")
        target = rel.get("Target")
        if not rid or not target:
            continue
        # OPC relationship Targets come in two forms: absolute from the package root ("/xl/
        # worksheets/sheet1.xml", which openpyxl writes) or relative to the referring part's
        # folder ("worksheets/sheet1.xml", relative to xl/_rels/ -> resolves under xl/). Normalize
        # both to the same "xl/worksheets/sheetN.xml" form used in the zip's own entry names.
        rid_to_target[rid] = target.lstrip("/") if target.startswith("/") else "xl/" + target

    try:
        workbook_xml = zf.read("xl/workbook.xml")
    except KeyError:
        return None
    root = ET.fromstring(workbook_xml)
    sheets_el = root.find(f"{{{_MAIN_NS}}}sheets")
    if sheets_el is None:
        return None

    for sheet_el in sheets_el:
        rid = sheet_el.get(f"{{{_REL_NS}}}id")
        target = rid_to_target.get(rid) if rid else None
        if target == target_part:
            return sheet_el.get("name")
    return None


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        # XlsxReader.kt has no special case for boolean cells: a <c t="b"> cell's raw <v> text is
        # literally "1" or "0", which is what falls out of the `else -> raw` branch there.
        return "1" if value else "0"
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        # Reverse openpyxl's date auto-conversion back to the bare Excel serial number the Kotlin
        # reader would have seen in <v>, so DateParse's serial fallback fires identically.
        serial = to_excel(value)
        if serial is None:
            return ""
        return _format_number(serial)
    if isinstance(value, (int, float)):
        return _format_number(value)
    return str(value)


def _format_number(value: float) -> str:
    if isinstance(value, int) or float(value).is_integer():
        return str(int(value))
    # Matches typical <v> text for a fractional numeric cell closely enough for downstream
    # parsing (paise/date extraction never need more than this); avoids Python float repr's
    # occasional long tails.
    return repr(value)


def read_rows(data: bytes) -> list[list[str]]:
    """Reads the lexicographically-first worksheet as a grid of cell strings.

    Blank cells become "", and rows that are entirely blank are dropped — matching
    ``XlsxReader.parseSheet``'s ``if (row.any { it.isNotBlank() }) rows.add(row)``.

    Raises on anything that isn't a readable xlsx (bad zip, missing workbook parts, corrupt XML) so
    the caller can distinguish "unreadable file" from "readable but empty file" the same way the
    Kotlin ``XlsxReader.read`` returning null does.
    """
    with zipfile.ZipFile(BytesIO(data)) as zf:
        sheet_name = _lexicographically_first_sheet_name(zf)

    workbook = openpyxl.load_workbook(BytesIO(data), data_only=True)
    try:
        if sheet_name is not None and sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.worksheets[0]

        rows: list[list[str]] = []
        for raw_row in worksheet.iter_rows(values_only=True):
            row = [_cell_to_str(v) for v in raw_row]
            if any(cell.strip() != "" for cell in row):
                rows.append(row)
        return rows
    finally:
        workbook.close()
