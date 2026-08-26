"""Tests for app/parsing/*.

These pin the same edge cases as the Kotlin test suite (DateParseTest.kt, NormalizeTest.kt,
CsvTest.kt, StatementParserTest.kt in app/src/test/) so behavior parity with the phone's parser is
actually verified, not assumed. Where a Python test's input/expected pair is lifted directly from
one of those files, the docstring/comment says so.
"""

import sys
import zipfile
from datetime import datetime
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo

# The server has no packaging/conftest.py setup, so make `app` importable when pytest is invoked
# straight from server/ (its default "prepend" import mode only puts tests/ itself on sys.path,
# not its parent).
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
import pytest

from app.parsing import csv_reader
from app.parsing.dates import parse_date, parse_paise
from app.parsing.normalize import display_merchant, normalize_merchant, vpa_handle
from app.parsing.statement import parse_statement

KOLKATA = ZoneInfo("Asia/Kolkata")


def _noon_ms(year: int, month: int, day: int) -> int:
    return int(datetime(year, month, day, 12, 0, tzinfo=KOLKATA).timestamp() * 1000)


# --- dates.py -----------------------------------------------------------------------------------


class TestParseDate:
    @pytest.mark.parametrize(
        "cell",
        [
            "08-05-2026", "08/05/2026", "8-5-26", "08-May-2026", "08-may-26",
            "2026-05-08", "08.05.2026",
        ],
    )
    def test_accepts_the_formats_indian_banks_actually_emit(self, cell):
        """DateParseTest.kt: `accepts the formats indian banks actually emit`."""
        assert parse_date(cell) == _noon_ms(2026, 5, 8), f"failed on {cell}"

    def test_ignores_a_trailing_time_component(self):
        """DateParseTest.kt: `ignores a trailing time component`."""
        assert parse_date("08/05/2026 14:32:11") == _noon_ms(2026, 5, 8)

    @pytest.mark.parametrize("cell", ["not a date", "", None])
    def test_rejects_junk(self, cell):
        """DateParseTest.kt: `rejects junk`."""
        assert parse_date(cell) is None

    def test_ddmmmyy_and_ddmmmyyyy_no_separator_formats(self):
        """The two no-separator formats (ddMMMyy, ddMMMyyyy) aren't in the pinned Kotlin test's
        sample list, but are in FORMATS, so exercise them directly."""
        assert parse_date("08May26") == _noon_ms(2026, 5, 8)
        assert parse_date("08May2026") == _noon_ms(2026, 5, 8)

    def test_bare_two_digit_year_bases_at_2000_not_a_rolling_pivot(self):
        assert parse_date("1-1-99") == _noon_ms(2099, 1, 1)

    def test_excel_serial_fallback_for_bare_integers(self):
        # 45786 is 2025-05-09 under the 1899-12-30 epoch.
        assert parse_date("45786") == _noon_ms(2025, 5, 9)

    def test_excel_serial_out_of_plausible_range_is_rejected(self):
        assert parse_date("5") is None
        assert parse_date("999999") is None


class TestParsePaise:
    def test_parses_rupee_strings_to_paise(self):
        """DateParseTest.kt: `parses rupee strings to paise`."""
        assert parse_paise("100.00") == 10_000
        assert parse_paise("1,23,456.78") == 1_23_456_78
        assert parse_paise("Rs. 500") == 50_000
        assert parse_paise("INR 899.00") == 89_900
        assert parse_paise("₹100") == 10_000

    def test_negatives_are_preserved_in_every_notation_banks_use(self):
        """DateParseTest.kt: `negatives are preserved in every notation banks use`."""
        assert parse_paise("-100.00") == -10_000
        assert parse_paise("(100.00)") == -10_000
        assert parse_paise("100.00 Dr") == -10_000
        assert parse_paise("100.00 Cr") == 10_000

    @pytest.mark.parametrize("cell", ["", "  ", "-", None])
    def test_blank_and_placeholder_cells_are_not_zero_amounts(self, cell):
        """DateParseTest.kt: `blank and placeholder cells are not zero amounts`."""
        assert parse_paise(cell) is None

    def test_dr_glued_to_the_number_has_no_word_boundary_and_is_left_unparsed(self):
        """`\\b[dc]r\\b` requires a word boundary before "Dr"; between a digit and a letter there
        isn't one, so a Dr glued directly onto the number (no space) doesn't get stripped by the
        regex, and the leftover letters make it an unparseable number, not a negative one. This
        mirrors the Kotlin regex exactly rather than being independently "smart" about it."""
        assert parse_paise("100.00Dr") is None


# --- normalize.py ---------------------------------------------------------------------------------


class TestNormalizeMerchant:
    def test_spelling_variants_of_one_merchant_collapse_to_the_same_key(self):
        """NormalizeTest.kt: `spelling variants of one merchant collapse to the same key`."""
        variants = ["SWIGGY", "Swiggy", "SWIGGY LIMITED", "Swiggy Pvt Ltd", "swiggy@ybl", "swiggy  "]
        assert {normalize_merchant(v) for v in variants} == {"swiggy"}

    def test_corporate_suffixes_are_stripped_repeatedly(self):
        """NormalizeTest.kt: `corporate suffixes are stripped repeatedly`."""
        assert normalize_merchant("Amazon Pay India Pvt Ltd") == "amazon pay"
        assert normalize_merchant("RELIANCE RETAIL LIMITED") == "reliance retail"

    def test_vpa_handle_is_excluded_from_the_merchant_key_but_kept_separately(self):
        """NormalizeTest.kt: `vpa handle is excluded from the merchant key but kept separately`."""
        assert normalize_merchant("bluetokai@okhdfcbank") == "bluetokai"
        assert normalize_merchant("bluetokai@ybl") == "bluetokai"
        assert vpa_handle("bluetokai@okhdfcbank") == "okhdfcbank"
        assert vpa_handle("bluetokai@ybl") == "ybl"
        assert vpa_handle("SWIGGY") is None

    def test_trailing_order_numbers_are_dropped(self):
        """NormalizeTest.kt: `trailing order numbers are dropped`."""
        assert normalize_merchant("ZEPTO*Order 8812") == "zepto order"

    def test_numeric_vpa_is_preserved(self):
        """NormalizeTest.kt: `numeric vpa is preserved`."""
        assert normalize_merchant("9876543210@ybl") == "9876543210"

    def test_honorifics_are_stripped_from_personal_payees(self):
        """NormalizeTest.kt: `honorifics are stripped from personal payees`."""
        assert normalize_merchant("MR DEEPAK KUMAR") == "deepak kumar"
        assert normalize_merchant("Deepak Kumar") == "deepak kumar"
        assert normalize_merchant("SMT ANITA RAO") == "anita rao"

    def test_a_bare_honorific_is_not_stripped_away_to_nothing(self):
        """NormalizeTest.kt: `a bare honorific is not stripped away to nothing`."""
        assert normalize_merchant("MR") == "mr"

    def test_blank_input_yields_an_empty_key(self):
        """NormalizeTest.kt: `blank input yields an empty key`."""
        assert normalize_merchant(None) == ""
        assert normalize_merchant("   ") == ""

    def test_distinct_merchants_stay_distinct(self):
        """NormalizeTest.kt: `distinct merchants stay distinct`."""
        assert normalize_merchant("ZOMATO") != normalize_merchant("SWIGGY")

    def test_display_casing_is_for_humans_only(self):
        """NormalizeTest.kt: `display casing is for humans only`."""
        assert display_merchant("BLUE TOKAI COFFEE") == "Blue Tokai Coffee"


# --- csv_reader.py --------------------------------------------------------------------------------


class TestCsvReader:
    def test_parses_simple_rows(self):
        """CsvTest.kt: `parses simple rows`."""
        assert csv_reader.read_rows("a,b,c\n1,2,3") == [["a", "b", "c"], ["1", "2", "3"]]

    def test_quoted_field_keeps_embedded_delimiter(self):
        """CsvTest.kt: `quoted field keeps embedded delimiter`."""
        rows = csv_reader.read_rows('date,"PAYMENT TO ACME, MUMBAI",100')
        assert rows[0] == ["date", "PAYMENT TO ACME, MUMBAI", "100"]

    def test_escaped_quotes_are_unescaped(self):
        """CsvTest.kt: `escaped quotes are unescaped`."""
        rows = csv_reader.read_rows('"he said ""hi""",2')
        assert rows[0] == ['he said "hi"', "2"]

    def test_quoted_field_may_span_newlines(self):
        """CsvTest.kt: `quoted field may span newlines`."""
        rows = csv_reader.read_rows('a,"line one\nline two",c')
        assert len(rows[0]) == 3
        assert rows[0][1] == "line one\nline two"

    def test_crlf_is_one_row_terminator(self):
        """CsvTest.kt: `crlf is one row terminator`."""
        assert len(csv_reader.read_rows("a,b\r\nc,d")) == 2

    def test_blank_rows_are_dropped(self):
        """CsvTest.kt: `blank rows are dropped`."""
        assert len(csv_reader.read_rows("a,b\n\n\n,,\nc,d")) == 2

    def test_detects_tab_and_semicolon_delimited_files(self):
        """CsvTest.kt: `detects tab and semicolon delimited files`."""
        assert csv_reader.detect_delimiter("a\tb\tc\n1\t2\t3") == "\t"
        assert csv_reader.detect_delimiter("a;b;c\n1;2;3") == ";"
        assert csv_reader.detect_delimiter("a,b,c\n1,2,3") == ","

    def test_preamble_line_does_not_confuse_delimiter_detection(self):
        """CsvTest.kt: `preamble line does not confuse delimiter detection`."""
        text = "Statement of account\nDate,Narration,Amount\n01/05/2026,X,10"
        assert csv_reader.detect_delimiter(text) == ","

    def test_bom_is_stripped(self):
        rows = csv_reader.read_rows("﻿a,b\n1,2")
        assert rows[0] == ["a", "b"]


# --- statement.py: synthetic CSV scenario ---------------------------------------------------------

HDFC_STYLE_CSV = """Account Statement
Account No: XXXXXXXX1234
Period: 01/05/2026 to 31/05/2026

Value Date,Narration,Ref No,Withdrawal Amt,Deposit Amt,Closing Balance
02/05/2026,UPI/512345678901/PAYMENT TO ARJUN MEHTA,512345678901,500.00,,9500.00
05/05/2026,UPI/512345678906/PAYMENT TO NETFLIX COM,,499.00,,9001.00
07/05/2026,UPI/512345678903/FROM RAVI KUMAR,,,10000.00,19001.00
08/05/2026,NIL ENTRY,,0.00,,19001.00
"""


class TestParseStatementCsv:
    def test_finds_header_row_beneath_preamble_and_parses_rows(self):
        result = parse_statement(HDFC_STYLE_CSV.encode("utf-8"))
        assert result.problem is None
        assert len(result.rows) == 3
        assert result.skipped == 1  # the zero-amount NIL ENTRY row
        assert result.total_data_rows == 4

    def test_debit_and_credit_columns_set_direction_and_amount(self):
        rows = parse_statement(HDFC_STYLE_CSV.encode("utf-8")).rows
        assert rows[0].direction == "DEBIT"
        assert rows[0].amount_paise == 50_000
        assert rows[0].balance_paise == 9_50_000
        assert rows[2].direction == "CREDIT"
        assert rows[2].amount_paise == 10_00_000

    def test_reference_column_used_when_present(self):
        rows = parse_statement(HDFC_STYLE_CSV.encode("utf-8")).rows
        assert rows[0].reference == "512345678901"

    def test_reference_falls_back_to_narration_regex_when_column_blank(self):
        """Row 2 (NETFLIX) has an empty Ref No cell; its UPI reference only exists embedded in the
        narration and must be extracted from there."""
        rows = parse_statement(HDFC_STYLE_CSV.encode("utf-8")).rows
        assert rows[1].reference == "512345678906"

    def test_merchant_extraction_and_channel(self):
        rows = parse_statement(HDFC_STYLE_CSV.encode("utf-8")).rows
        assert rows[0].merchant_raw == "ARJUN MEHTA"
        assert rows[0].merchant_norm == "arjun mehta"
        assert rows[0].channel == "UPI"

    def test_occurred_at_is_noon_asia_kolkata(self):
        rows = parse_statement(HDFC_STYLE_CSV.encode("utf-8")).rows
        assert rows[0].occurred_at_ms == _noon_ms(2026, 5, 2)


# --- statement.py: the same scenario via a real .xlsx file ----------------------------------------


def _build_xlsx_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Account Statement"])
    ws.append(["Account No: XXXXXXXX1234"])
    ws.append([])
    ws.append(["Value Date", "Narration", "Ref No", "Withdrawal Amt", "Deposit Amt", "Closing Balance"])
    ws.append(["02/05/2026", "UPI/512345678901/PAYMENT TO ARJUN MEHTA", "512345678901", 500.00, None, 9500.00])
    ws.append(["05/05/2026", "UPI/512345678906/PAYMENT TO NETFLIX COM", None, 499.00, None, 9001.00])
    ws.append(["07/05/2026", "UPI/512345678903/FROM RAVI KUMAR", None, None, 10000.00, 19001.00])
    ws.append(["08/05/2026", "NIL ENTRY", None, 0.00, None, 19001.00])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestParseStatementXlsx:
    """Exercises the real openpyxl read path (not a mock) via an in-memory workbook."""

    def test_parses_the_same_scenario_as_the_csv_test(self):
        result = parse_statement(_build_xlsx_bytes())
        assert result.problem is None
        assert len(result.rows) == 3
        assert result.skipped == 1

        rows = result.rows
        assert rows[0].direction == "DEBIT"
        assert rows[0].amount_paise == 50_000
        assert rows[0].reference == "512345678901"
        assert rows[0].merchant_norm == "arjun mehta"
        assert rows[1].reference == "512345678906"
        assert rows[2].direction == "CREDIT"
        assert rows[2].amount_paise == 10_00_000

    def test_picks_the_lexicographically_first_sheet_not_the_active_one(self):
        """openpyxl orders worksheets by tab position; the Kotlin reader picks whichever
        xl/worksheets/sheetN.xml part sorts first by filename. openpyxl's own writer always
        re-serializes sheet1.xml/sheet2.xml/... in final tab order, so a workbook built and saved
        by openpyxl alone can never exhibit the divergence being tested here — real Excel-produced
        files can, because Excel keeps a sheet's original part name stable across tab reorders.
        To exercise that case, build the workbook normally, then hand-edit the saved zip's
        xl/workbook.xml so tab position 0 points (via r:id) at sheet2.xml while sheet1.xml — still
        holding "FROM SHEET ONE FILE" — becomes tab position 1. This reproduces the exact
        structure a reordered real-world file has, without depending on openpyxl's writer doing
        the reordering for us."""
        wb = openpyxl.Workbook()
        first = wb.active
        first.title = "SheetA"
        first.append(["Date", "Narration", "Amount"])
        first.append(["08-05-2026", "FROM SHEET ONE FILE", -100.00])

        second = wb.create_sheet("SheetB")
        second.append(["Date", "Narration", "Amount"])
        second.append(["08-05-2026", "FROM SHEET TWO FILE", -200.00])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        with zipfile.ZipFile(buf) as zf:
            entries = {name: zf.read(name) for name in zf.namelist()}

        workbook_xml = entries["xl/workbook.xml"].decode("utf-8")
        assert 'name="SheetA"' in workbook_xml and 'r:id="rId1"' in workbook_xml
        # Swap the <sheet> element order so SheetB (rId2 -> sheet2.xml) is declared first, while
        # rId1 -> sheet1.xml (SheetA, "FROM SHEET ONE FILE") is declared second. Relationship IDs,
        # and therefore which r:id targets which physical sheetN.xml file, are untouched.
        sheet_a_el = '<sheet name="SheetA" sheetId="1" state="visible" r:id="rId1" />'
        sheet_b_el = '<sheet name="SheetB" sheetId="2" state="visible" r:id="rId2" />'
        assert sheet_a_el in workbook_xml and sheet_b_el in workbook_xml
        reordered = workbook_xml.replace(
            sheet_a_el + sheet_b_el, sheet_b_el + sheet_a_el
        )
        assert reordered != workbook_xml
        entries["xl/workbook.xml"] = reordered.encode("utf-8")

        out = BytesIO()
        with zipfile.ZipFile(out, "w") as zf:
            for name, content in entries.items():
                zf.writestr(name, content)

        result = parse_statement(out.getvalue())
        assert result.problem is None
        assert len(result.rows) == 1
        # sheet1.xml (SheetA) is lexicographically first by filename despite being tab position 1.
        assert result.rows[0].merchant_raw == "FROM SHEET ONE FILE"

    def test_date_formatted_cell_round_trips_through_excel_serial(self):
        """A cell openpyxl loads as a real python date (because Excel applied date formatting to
        it) must still parse, via the same Excel-serial fallback DateParse uses for a raw xlsx
        date cell — not by accident matching one of the text date formats."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "Narration", "Amount"])
        ws.append([__import__("datetime").date(2026, 5, 8), "ACME PURCHASE", -100.00])
        buf = BytesIO()
        wb.save(buf)

        result = parse_statement(buf.getvalue())
        assert result.problem is None
        assert len(result.rows) == 1
        assert result.rows[0].occurred_at_ms == _noon_ms(2026, 5, 8)


# --- statement.py: delimiter variants --------------------------------------------------------------


class TestParseStatementDelimiters:
    def test_semicolon_delimited_statement(self):
        text = (
            "Date;Particulars;Amount\n"
            "08-05-2026;POS PURCHASE ACME;250.00 Dr\n"
            "08-05-2026;REFUND ACME;250.00 Cr\n"
        )
        result = parse_statement(text.encode("utf-8"))
        assert result.problem is None
        assert len(result.rows) == 2
        assert result.rows[0].direction == "DEBIT"
        assert result.rows[1].direction == "CREDIT"

    def test_tab_delimited_statement_with_signed_amount_column(self):
        """StatementParserTest.kt: `single signed amount column encodes direction` (ICICI-style)."""
        text = (
            "Txn Date\tTransaction Remarks\tAmount\tBalance\n"
            "08-05-2026\tUPI/512345678902/VENDCO\t-99.00\t10261.13\n"
            "08-05-2026\tSALARY CREDIT\t50000.00\t60261.13\n"
        )
        result = parse_statement(text.encode("utf-8"))
        assert result.problem is None
        rows = result.rows
        assert len(rows) == 2
        assert rows[0].direction == "DEBIT"
        assert rows[0].amount_paise == 9_900
        assert rows[1].direction == "CREDIT"
        assert rows[1].amount_paise == 50_00_000


# --- statement.py: further pinned StatementParserTest.kt edge cases -------------------------------


class TestParseStatementEdgeCases:
    def test_specific_column_names_win_over_generic_ones(self):
        """StatementParserTest.kt: `specific column names win over generic ones` — "withdrawal
        amount" must not be beaten by a substring match on "amount"."""
        text = "Date,Narration,Withdrawal Amount,Deposit Amount\n08-05-2026,ACME,500.00,\n"
        result = parse_statement(text.encode("utf-8"))
        assert len(result.rows) == 1
        assert result.rows[0].direction == "DEBIT"
        assert result.rows[0].amount_paise == 50_000

    def test_unanchored_long_numbers_are_not_treated_as_references(self):
        """StatementParserTest.kt: `unanchored long numbers are not treated as references`."""
        text = "Date,Narration,Amount\n08-05-2026,TRANSFER TO 123456789012 ACME,-100.00\n"
        result = parse_statement(text.encode("utf-8"))
        assert result.rows[0].reference is None

    def test_rows_with_no_usable_date_or_amount_are_skipped_not_failed(self):
        """StatementParserTest.kt: `rows with no usable date or amount are skipped not failed`."""
        text = (
            "Date,Narration,Amount\n"
            "08-05-2026,GOOD ROW,-100.00\n"
            "TOTAL,,\n"
            ",,\n"
            "not-a-date,JUNK,-50.00\n"
        )
        result = parse_statement(text.encode("utf-8"))
        assert len(result.rows) == 1
        assert result.skipped == 2
        assert result.problem is None

    def test_zero_amount_rows_are_skipped(self):
        """StatementParserTest.kt: `zero amount rows are skipped`."""
        text = "Date,Narration,Amount\n08-05-2026,NIL ENTRY,0.00\n"
        assert parse_statement(text.encode("utf-8")).rows == []

    def test_reports_a_clear_problem_when_the_file_is_not_a_statement(self):
        """StatementParserTest.kt: `reports a clear problem when the file is not a statement`."""
        result = parse_statement(b"hello world\nthis is not a statement")
        assert result.problem is not None
        assert result.rows == []


class TestReferenceNormalization:
    """StatementParser.parseRows hands its resolved `ref` to the `txnRecord()` builder in
    TxnRecord.kt, whose refNo parameter is `RefKey.normalize(refNo)` — strip non-alphanumerics,
    uppercase, and require at least 6 characters with at least 4 digits, or null. This is what lets
    a reference embedded in a statement narration and the same reference normalized from an SMS
    collapse to the same dedupe key server-side, so it applies here too, not just to ref-column
    values that are already clean."""

    def test_punctuation_in_the_ref_column_is_stripped_and_uppercased(self):
        text = "Date,Narration,Amount,Ref No\n08-05-2026,ACME,-100.00,utr/512-345.678\n"
        result = parse_statement(text.encode("utf-8"))
        assert result.rows[0].reference == "UTR512345678"

    def test_a_reference_with_too_few_digits_is_rejected(self):
        """Mirrors the "UPI Mandate" incident in docs/ARCHITECTURE.md#deduplication: a
        reference-shaped string with fewer than 4 digits must not become a dedupe key."""
        text = "Date,Narration,Amount,Ref No\n08-05-2026,ACME,-100.00,MANDATE\n"
        result = parse_statement(text.encode("utf-8"))
        assert result.rows[0].reference is None

    def test_a_reference_shorter_than_six_characters_is_rejected(self):
        text = "Date,Narration,Amount,Ref No\n08-05-2026,ACME,-100.00,1234\n"
        result = parse_statement(text.encode("utf-8"))
        assert result.rows[0].reference is None

    def test_narration_extracted_reference_is_also_normalized(self):
        text = "Date,Narration,Amount\n08-05-2026,UPI/abc123def456/PAYMENT TO ACME,-100.00\n"
        result = parse_statement(text.encode("utf-8"))
        assert result.rows[0].reference == "ABC123DEF456"


# --- format sniffing / failure modes ---------------------------------------------------------------


class TestFormatSniffing:
    def test_legacy_xls_biff_magic_bytes_produce_a_problem_not_a_crash(self):
        data = bytes([0xD0, 0xCF, 0x11, 0xE0, 0x00, 0x00, 0x00, 0x00]) + b"\x00" * 20
        result = parse_statement(data)
        assert result.rows == []
        assert result.problem is not None
        assert "xls" in result.problem.lower()

    def test_empty_file_produces_a_problem_not_a_crash(self):
        result = parse_statement(b"")
        assert result.rows == []
        assert result.problem is not None

    def test_garbage_binary_that_is_not_a_zip_or_biff_file_produces_a_problem_not_a_crash(self):
        data = bytes(range(256)) * 4
        result = parse_statement(data)
        assert result.rows == []
        assert result.problem is not None

    def test_corrupt_zip_masquerading_as_xlsx_produces_a_problem_not_a_crash(self):
        data = b"PK\x03\x04" + b"not actually a valid zip central directory" * 3
        result = parse_statement(data)
        assert result.rows == []
        assert result.problem is not None


# --- regression: SBI's real narration format (found via a real statement upload) ------------------
#
# SBI inserts a direction/type tag between the UPI marker and the reference ("UPI/DR/<ref>/..."),
# and appends a fixed teller/branch suffix to every narration ("<acct-ref> AT <code> <branch>").
# Both defeated the original marker-then-digits reference regex and the merchant letter-count
# heuristic respectively — the reference bug is the serious one: it silently made most rows share
# one of a handful of fake "reference" values, which would collide against the server's global
# unique reference index and drop real transactions as false duplicates on import.

SBI_STYLE_CSV = (
    "Date,Details,Ref No/Cheque No,Debit,Credit,Balance\n"
    "01/04/2026, WDL TFR   UPI/DR/609195951332/SBI card/KKBK/sbicardp.b/Pay   0097692162094 AT 11309 PILANI,,2025.00,,28447.04\n"
    "01/04/2026, WDL TFR   UPI/DR/609193089121/SBI card/KKBK/sbicardp.b/Pay   0097692162094 AT 11309 PILANI,,2025.00,,26422.04\n"
    "01/04/2026, WDL TFR   UPI/DR/609102594284/ZEPTO/HDFC/zepto.payu/UPI   0097692162094 AT 11309 PILANI,,105.00,,26317.04\n"
    "03/04/2026, WDL TFR   UPI/DR/109515786107/MOHAMMED/ICIC/nasirshaik/na   0097694162092 AT 11309 PILANI,,1100.00,,94895.04\n"
    "02/04/2026, DEP TFR   NEFT*IDFB0010201*IDFBH26092821888*RZPX PVT LTD PA   0099509044300 AT 11309 PILANI,,,68600.00,94610.04\n"
)


class TestSbiNarrationFormat:
    def test_reference_extracted_despite_dr_tag_between_marker_and_digits(self):
        rows = parse_statement(SBI_STYLE_CSV.encode("utf-8")).rows
        assert rows[0].reference == "609195951332"
        assert rows[1].reference == "609193089121"
        assert rows[2].reference == "609102594284"

    def test_distinct_transactions_get_distinct_references_not_the_shared_branch_suffix(self):
        """The bug: every row here shares the same trailing '0097692162094 AT 11309 PILANI'
        boilerplate. Before the fix, all four UPI rows extracted that shared number as their
        "reference" instead of the real per-transaction one right after UPI/DR/."""
        rows = parse_statement(SBI_STYLE_CSV.encode("utf-8")).rows
        refs = [r.reference for r in rows if r.reference]
        assert len(refs) == len(set(refs)), f"references collided: {refs}"
        assert "0097692162094" not in refs

    def test_neft_star_delimited_reference_is_not_misextracted(self):
        """NEFT*<ifsc>*<utr>*<name> has no reliable single-token reference adjacent to the marker
        (the IFSC sits between NEFT and the real UTR) — no reference is safer than a wrong one."""
        rows = parse_statement(SBI_STYLE_CSV.encode("utf-8")).rows
        assert rows[4].reference is None

    def test_merchant_is_not_the_trailing_branch_suffix(self):
        rows = parse_statement(SBI_STYLE_CSV.encode("utf-8")).rows
        for row in rows:
            assert row.merchant_raw is not None
            assert "PILANI" not in row.merchant_raw
            assert "0097692162094" not in row.merchant_raw
            assert "0097694162092" not in row.merchant_raw

    def test_merchant_is_not_the_wdl_tfr_upi_prefix(self):
        """"WDL TFR   UPI" out-letter-counts the real merchant segment purely by being a 3-word
        compound — none of its words are boilerplate on their own, so whole-segment scheme-word
        matching doesn't catch it; only per-word stripping does."""
        rows = parse_statement(SBI_STYLE_CSV.encode("utf-8")).rows
        for row in rows:
            assert row.merchant_raw.strip().upper() != "WDL TFR   UPI"
